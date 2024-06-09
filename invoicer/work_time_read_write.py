import logging
from datetime import date, datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def copy_time_report_txt_to_list(source_path):
    try:
        with open(source_path, 'r') as work_time_file:
            return work_time_file.readlines()
    except FileNotFoundError as e:
        logging.exception(e)


def get_last_day_of_current_month():
    current_date_for_today = datetime.today()
    my_next_month = current_date_for_today.replace(day=28) + timedelta(days=4)
    my_last_day = my_next_month - timedelta(days=my_next_month.day)
    return my_last_day.day


def get_int_day(str_date):
    return int(str_date[0:2])


def int_day_to_string(int_day):
    if len(str(int_day)) == 1:
        return str(f"0{int_day}")
    else:
        return str(int_day)


def int_month_to_string():
    int_month = date.today().month
    if len(str(int_month)) == 1:
        return str(f"0{int_month}")
    else:
        return str(int_month)


def updated_list_to_dict(lines_list):
    last_day = get_last_day_of_current_month()
    lines_dict = {}

    if (lines_list[0])[0:2] != "01":
        lines_list.insert(0, f"01.{int_month_to_string()}:")

    for line in lines_list:
        if lines_list.index(line) != len(lines_list) - 1:
            next_line = lines_list[lines_list.index(line) + 1]
            if get_int_day(line) + 1 != get_int_day(next_line):
                lines_list.insert(lines_list.index(line) + 1, f"{int_day_to_string(get_int_day(line) + 1)}"
                                                              f".{int_month_to_string()}:")
        if lines_list[lines_list.index(line)] == lines_list[-1] and (lines_list[-1])[0:2] != str(last_day):
            lines_list.insert(lines_list.index(line) + 1, f"{int_day_to_string(get_int_day(line) + 1)}"
                                                          f".{int_month_to_string()}:")
        if line[-1] == '\n':
            lines_list[lines_list.index(line)] = line[:-1]

        date_and_activity = line.split(":")
        lines_dict.update({f"{date_and_activity[0]}.{date.today().year}": date_and_activity[1]})

    return lines_dict


def save_dict_in_result(date_activity_dict, save_path):
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = "WorkTimeReport"

    table_headings = ["Date", "Name and surname", "Client", "Standard Time", "Overtime", "Task description"]

    work_sheet.append(table_headings)

    for date_key, activity_value in date_activity_dict.items():
        if not activity_value:
            row_of_table = [f"{date_key}", "John Doe", "TestClient", 0, 0, activity_value]
        else:
            row_of_table = [f"{date_key}", "John Doe", "TestClient", 8, 0, activity_value]
        work_sheet.append(row_of_table)

    work_sheet[f"D{work_sheet.max_row + 1}"] = f"=SUM(D1:D{work_sheet.max_row})"

    for line in range(2, work_sheet.max_row):
        if work_sheet[f"D{line}"].value == 0:
            for column in range(1, 7):
                work_sheet[f"{get_column_letter(column)}{line}"].fill = PatternFill(fgColor="699949",
                                                                                    fill_type="solid")

    work_book.save(f"{save_path}.xlsx")


def create_work_report_from_txt(txt_path, save_location):
    list_of_lines = copy_time_report_txt_to_list(txt_path)
    dict_of_line = updated_list_to_dict(list_of_lines)
    save_dict_in_result(dict_of_line, save_location)


def get_amount_of_working_hours(save_location):
    try:
        work_book = load_workbook(save_location)
        work_sheet = work_book.active

        hours_list = [work_sheet[f"D{line}"].value for line in range(2, work_sheet.max_row)]
        hours_sum = sum(hours_list)

        print(hours_sum)
        return hours_sum
    except FileNotFoundError as e:
        logging.exception(e)
